import mysql.connector
from tkinter import *
from tkinter import ttk, messagebox, simpledialog
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from datetime import datetime

# MySQL configuration
DB_CONFIG = {
    'host': 'localhost',
    'user': 'root',
    'password': 'harsh@2002',
    'database': 'attendance_system'
}

def fetch_data(name=None, department=None, specific_date=None):
    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        cursor = conn.cursor()

        query = """
        SELECT s.student_id, s.name, s.department, s.email, a.date, a.time
        FROM students s
        JOIN attendance a ON s.student_id = a.student_id
        WHERE 1=1
        """
        params = []

        if name:
            query += " AND s.name LIKE %s"
            params.append(f"%{name}%")
        if department:
            query += " AND s.department LIKE %s"
            params.append(f"%{department}%")
        if specific_date:
            try:
                datetime.strptime(specific_date, '%Y-%m-%d')
                query += " AND a.date = %s"
                params.append(specific_date)
            except ValueError:
                messagebox.showerror("Invalid Date", "Enter date as YYYY-MM-DD")
                return []

        query += " ORDER BY a.date DESC, a.time DESC"
        cursor.execute(query, params)
        return cursor.fetchall()

    except mysql.connector.Error as err:
        messagebox.showerror("DB Error", str(err))
        return []
    finally:
        if cursor:
            cursor.close()
        if conn and conn.is_connected():
            conn.close()

def populate_table(data):
    for row in tree.get_children():
        tree.delete(row)
    for row in data:
        tree.insert("", END, values=row)

def initial_load():
    data = fetch_data()
    populate_table(data)

def search_data():
    name = name_entry.get().strip()
    dept = dept_entry.get().strip()
    specific_date = date_entry.get().strip()
    data = fetch_data(name, dept, specific_date)
    populate_table(data)

def export_excel():
    name = name_entry.get().strip()
    dept = dept_entry.get().strip()
    specific_date = date_entry.get().strip()
    data = fetch_data(name, dept, specific_date)

    if not data:
        messagebox.showinfo("No Data", "No data to export.")
        return

    file_name = simpledialog.askstring("Filename", "Enter Excel file name (without extension):")
    if not file_name:
        return

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Report"

        headers = ["Student ID", "Name", "Department", "Email", "Date", "Time"]
        ws.append(headers)

        bold = Font(bold=True)
        for col_num, col_name in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = bold

        for row in data:
            ws.append(row)

        for col_num in range(1, len(headers) + 1):
            max_len = max(len(str(ws.cell(row=i, column=col_num).value)) for i in range(1, len(data) + 2))
            ws.column_dimensions[get_column_letter(col_num)].width = max_len + 2

        wb.save(f"{file_name}.xlsx")
        messagebox.showinfo("Success", f"Data saved to {file_name}.xlsx")

    except Exception as e:
        messagebox.showerror("Export Error", str(e))

# GUI Setup
root = Tk()
root.title("Attendance Viewer & Exporter")

# Filters
Label(root, text="Student Name:").grid(row=0, column=0, padx=5, pady=5, sticky=E)
name_entry = Entry(root)
name_entry.grid(row=0, column=1, padx=5, pady=5)

Label(root, text="Department:").grid(row=0, column=2, padx=5, pady=5, sticky=E)
dept_entry = Entry(root)
dept_entry.grid(row=0, column=3, padx=5, pady=5)

Label(root, text="Date (YYYY-MM-DD):").grid(row=0, column=4, padx=5, pady=5, sticky=E)
date_entry = Entry(root)
date_entry.grid(row=0, column=5, padx=5, pady=5)

Button(root, text="Search", command=search_data).grid(row=0, column=6, padx=10)
Button(root, text="Download Excel", command=export_excel).grid(row=1, column=6, pady=10)

# Table
columns = ("Student ID", "Name", "Department", "Email", "Date", "Time")
tree = ttk.Treeview(root, columns=columns, show="headings")
for col in columns:
    tree.heading(col, text=col)
    tree.column(col, anchor=CENTER)
tree.grid(row=2, column=0, columnspan=7, padx=10, pady=10)

initial_load()
root.mainloop()
